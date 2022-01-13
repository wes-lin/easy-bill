import json
import glob
import shutil
import openpyxl
import datetime
import os
import logging
from jinja2 import Template
from jinja2.filters import FILTERS
from CustomerFilter import datetimeformat
from ImageBill import ImageBill
from ImageOCR import ImageOCR
from PDFBill import PDFBill
import ChromeDriverDownload

FILTERS['datetimeformat'] = datetimeformat
logging.basicConfig(encoding='utf-8',format='%(asctime)s %(message)s')

def load_config(path):
    logging.warning('加载配置文件')
    with open(path,mode='r',encoding="utf-8") as configFile:
        config = json.load(configFile)
        logging.warning(config)
    return config

def process(bill,config,taget):
    tmp = {}
    tmp.update(config)
    tmp.update(bill)
    bill['outPath'] = Template(taget['outPath']).render(**tmp)
    bill['outFileName'] = Template(taget['outFileName']).render(**tmp)
    folder = os.path.exists(bill['outPath'])
    if not folder:
        os.makedirs(bill['outPath'])
    shutil.copy(bill['sourcePath'], bill['outPath']+'/'+bill['outFileName']+bill['ext'])
    logging.warning(bill)

def outExcel(config,array):
    outPath = config['outPath']
    wb = openpyxl.Workbook()
    for index,item in enumerate(array):
        name = item['name']
        bills = item['data']
        outExcel = item['outExcel']
        ws = wb.create_sheet(title=name,index=index)
        for rindex , bill in enumerate(bills):
            tmp = {}
            tmp.update(config)
            tmp.update(bill)
            for cindex,column in enumerate(outExcel): 
                c = ws.cell(row=(rindex+1),column=(cindex+1))
                if 'numformat' in column:
                    c.number_format = column['numformat']
                v = tmp[column['title']]
                if 'type' in column:
                    if column['type'] == 'float':
                        v = float(v)
                    if column['type'] == 'date':
                        v = datetime.datetime.strptime(v,'%Y-%m-%d')
                c.value = v
    wb.save(outPath+'/outExcel.xlsx')
    wb.close()

if __name__=='__main__':
    config = load_config('../config.json')
    outExcelData = []
    for taget in config['taget']:
        type = taget['type']
        if 'disable' in taget and taget['disable']:
            continue
        bills = []
        if type == 'pdf':
            paths = glob.glob(config['source']+'/*.pdf')
            didiTravel = {}
            for path in paths:
                bill = PDFBill(path)
                bill.extract()
                data = bill.getData()
                #记录滴滴行程单订单日期，为滴滴电子发票记录订单日期
                if bill.isDiDiTravel():
                    didiTravel[data['applyDate']+'_'+data['money']] = data['orderDate']
                bills.append(data)

            for bill in bills:
                key = bill['applyDate']+'_'+bill['money']
                if 'orderDate' not in bill: 
                    if key in didiTravel:
                        bill['orderDate'] = didiTravel[key]
                    else:
                        bill['orderDate'] = bill['applyDate']
                process(bill,config,taget)
        if type =='imgae':
            ChromeDriverDownload.update()
            paths = glob.glob(config['source']+'/*[.jpg, .png]')
            ocr = ImageOCR(config['debug'])
            texts = ocr.extract(paths)
            ocr.quit()
            for index , text in enumerate(texts):
                bill = ImageBill(paths[index],text)
                bill.extract()
                data = bill.getData()
                bills.append(data)
                process(data,config,taget)
        
        outExcelData.append({
                'name':taget['name'],
                'data':bills,
                'outExcel':taget['outExcel']
            })

    outExcel(config,outExcelData)
